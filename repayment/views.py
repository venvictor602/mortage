

import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
import os
from django.conf import settings

# Path to the template file
TEMPLATE_FILE_PATH = 'repayment/files/our template.xlsx'

def process_repayment(request):
    if request.method == 'POST':
        # Retrieve form data
        selected_month = request.POST.get('month')
        selected_year = request.POST.get('year')
        akpab_file = request.FILES.get('akpab_file')

        if not selected_month or not selected_year or not akpab_file:
            return HttpResponse("Please fill in all fields and upload the required file.", status=400)

        # Get the last two digits of the year
        last_two_digits_year = str(selected_year)[-2:]

        try:
            # Load AKPAB workbook
            akpab_wb = openpyxl.load_workbook(akpab_file)
            akpab_ws = akpab_wb.active

            # Load the template workbook
            template_wb = openpyxl.load_workbook(TEMPLATE_FILE_PATH)
            template_ws = template_wb.active

            # Define AKPAB column indices
            akpab_employee_number_col = 3  # EMPLOYMENT NUMBER is in column 3
            akpab_amount_col = 6            # AMOUNT is in column 6

            # Find "TOTAL" column in the template
            total_column = None
            header_row = 3  # Assuming the template headers are in row 3

            for col in range(1, template_ws.max_column + 1):
                header_value = template_ws.cell(row=header_row, column=col).value
                if header_value == 'TOTAL':  # Find the 'TOTAL' column
                    total_column = col
                    break

            if total_column is None:
                return HttpResponse("Error: 'TOTAL' column not found in the template.", status=500)

            # Create a new header for the current month/year before the "TOTAL" column
            new_month_year_column_name = f"{selected_month} {last_two_digits_year}"
            template_ws.insert_cols(total_column)  # Insert new column before the "TOTAL"
            template_ws.cell(row=header_row, column=total_column, value=new_month_year_column_name)

            # Map rows from AKPAB to the template based on employee number
            for akpab_row in akpab_ws.iter_rows(min_row=4):  # Start processing from row 4
                akpab_employee_number = akpab_row[akpab_employee_number_col - 1].value
                akpab_amount = akpab_row[akpab_amount_col - 1].value

                # Check if the EMPLOYEE NUMBER is numeric
                if akpab_employee_number is not None and str(akpab_employee_number).strip().isdigit():
                    akpab_employee_number = int(akpab_employee_number)  # Convert to integer

                    # Proceed only if the amount is numeric
                    if akpab_amount is not None and isinstance(akpab_amount, (int, float)):
                        # Loop through template rows to match the employee number
                        for template_row in range(4, template_ws.max_row + 1):  # Start from row 4
                            template_employee_number = template_ws.cell(row=template_row, column=3).value  # EMP_CODE

                            # Match employee number from AKPAB with EMP_CODE in the template
                            if template_employee_number == akpab_employee_number:
                                # Add the amount in the new month/year column
                                template_ws.cell(row=template_row, column=total_column, value=akpab_amount)

                                # Update total (now shifted one column right)
                                total_cell = template_ws.cell(row=template_row, column=total_column + 1)
                                total_cell.value = (total_cell.value or 0) + akpab_amount
                                break
                    else:
                        print(f"Skipping row {akpab_row[0].row}: Non-numeric amount '{akpab_amount}'.")
                else:
                    print(f"Skipping row {akpab_row[0].row}: Non-numeric employee number '{akpab_employee_number}'.")

            # Save the modified template
            template_wb.save(TEMPLATE_FILE_PATH)

            # Prepare the filename using the selected month and last two digits of the year
            downloaded_filename = f"updated_template_{selected_month}_{last_two_digits_year}.xlsx"
            
            # Return the updated file as a response
            with open(TEMPLATE_FILE_PATH, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={downloaded_filename}'
                return response

        except Exception as e:
            return HttpResponse(f"Error processing the file: {str(e)}", status=500)

    # If it's a GET request, render your custom HTML form
    return render(request, 'index.html')





from django.http import FileResponse, HttpResponse
import os
from django.conf import settings

def download_template(request):
    # Define the template file path
    # TEMPLATE_FILE_PATH = os.path.join(settings.BASE_DIR, 'repayment', 'files', 'our_template.xlsx')

    # Log the path for debugging
    print("Looking for template file at:", TEMPLATE_FILE_PATH)

    # Check if the template file exists
    if os.path.exists(TEMPLATE_FILE_PATH):
        # Use FileResponse to return the file
        return FileResponse(open(TEMPLATE_FILE_PATH, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        # Log the issue or provide more detail in the response
        return HttpResponse(f"Template file not found at {TEMPLATE_FILE_PATH}.", status=404)
